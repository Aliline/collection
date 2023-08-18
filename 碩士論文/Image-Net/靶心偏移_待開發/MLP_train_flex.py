# 6.此為用於訓練keras模型 
import os
os.environ['CUDA_VISIBLE_DEVICES'] = '0'
from tkinter import X
from imblearn.over_sampling import SMOTE
import numpy as np
import pandas as pd
import multiprocessing
import numpy as np
import math
#from pandas.core.construction import array
from concurrent.futures import ThreadPoolExecutor
from PIL import Image  
from sklearn.model_selection import train_test_split
from alive_progress import alive_bar
#keras-------
import keras
from keras.layers import Input,Conv2D,MaxPooling2D,Dense,Flatten,concatenate,Dropout,BatchNormalization,LSTM
from keras.models import Sequential,Model
import numpy as np
from sklearn.metrics import confusion_matrix,classification_report,accuracy_score,mean_squared_error
from openpyxl import Workbook
import tensorflow as tf
from keras.utils import to_categorical
from keras.callbacks import EarlyStopping,ModelCheckpoint
from keras import regularizers
from keras.backend import set_session
from keras.backend import clear_session
from keras.backend import get_session
import gc
tf.config.run_functions_eagerly(True)
# CUDA_VISIBLE_DEVICES=0

flax = 28
savepath = "D:/F111156114_Aliline/imgtoword/wordtovector/測試/靶心偏移_待開發/flaxed_model/"
testpath = "D:/F111156114_Aliline/imgtoword/wordtovector/測試/靶心偏移_待開發/flaxed_model/"
#xy訓練資料
xvector = np.load('D:/F111156114_Aliline/imgtoword/wordtovector/vector_MLP_renew/vx260_full_corp.npy',allow_pickle=True)
yvector = np.load('D:/F111156114_Aliline/imgtoword/wordtovector/vector_MLP_renew/v260y_full_corp_flax{}.npy'.format(flax))
# xvector = np.load('D:/F111156114_Aliline/imgtoword/wordtovector/組合包/XY資料生成/x.npy',allow_pickle=True)
# yvector = np.load('D:/F111156114_Aliline/imgtoword/wordtovector/組合包/XY資料生成/y.npy')
testtxt = "D:/F111156114_Aliline/imgtoword/wordtovector/vector_txt/vectors_all_ver_.txt"
#文字向量y
testvecotr = np.load("D:/F111156114_Aliline/imgtoword/wordtovector/vectors_260_ver2.npy")
actiname = 'tanh'
# actiname = 'LeakyReLU'
# lr=0.09
lr=0.01
beta_1=0.9
beta_2=0.999
epsilon=1e-08
decay=0.0
ver = 0
ver_num = 0
#每Xepochs存一次檢查點
peroid = 250
#classes指詞遷入向量大小
classes=260
while True:
    if not os.path.exists(f'{savepath}MLP_{classes}_ver{ver_num}_{actiname}_{peroid}.h5'):
        ver = ver_num
        break
    ver_num = ver_num +1

def cul(testvecotr,y):
    vtotal = 0
    # for v in range(classes):
    #     vc = (y[v] - testvecotr[v])**2 #距離
    #     vtotal = vtotal + vc
        # temp = (testvecotr[v] - y[v])**2
    # vc = np.mean((y - testvecotr)**2)  #MSE

    temp = (y - testvecotr)**2
    temp = np.sum(temp)
    temp = np.sqrt(temp)
    # vtotal = math.sqrt(vtotal)
    # temp2 = mean_squared_error(testvecotr , y)

    # return temp2
    return temp

# Reset Keras Session
#無視
def mean_pred(y_true, y_pred):
    y_ture_L= []
    y_prob_L = []
    tempy_ture = y_true.numpy()
    total = len(tempy_ture)
    check = 0
    with alive_bar(total) as bar:
        for yt in y_true:
            ytnum = yt.numpy()
            for k in tf.range(1000):
                vtotal = np.float64(0)
                if (ytnum == testvecotr[k]).all() :
                    y_ture_L.append(k)
                    check = k
                    break
                for v in tf.range(100):
                    vc = ytnum[v] - testvecotr[k][v]
                    vtotal = vtotal + abs(vc)
                # print(vtotal.numpy())
                vtemp = vtotal.numpy()
                print("vtotal:{}".format(vtemp))
                if vtemp == 0.0:
                    y_ture_L.append(k)
                    check = k
                    break
            print("check:L")
            print(y_ture_L[check])
            bar()
    for y in y_pred:
            min = 0
            min_lab = -1
            for k in tf.range(1000):
                vtotal = np.float64(0)
                for v in tf.range(100):
                    # vc = y[v] - testvecotr[k][v] #距離
                    ynum = y.numpy()
                    vc = np.mean((testvecotr[k][v] - ynum[v])**2)  #MSE
                    vtotal = vtotal + vc
                    #改成MSE的計算公式
                    #Top5正確率試試看(距離前5短)
                if min_lab == -1 :
                    min = vtotal
                    min_lab = k
                elif min > vtotal:
                    min = vtotal
                    min_lab = k
            y_prob_L.append(min_lab)            
    Accuracy=accuracy_score(y_ture_L,y_prob_L)    
    return Accuracy

def reset_keras():
    sess = get_session()
    clear_session()
    sess.close()
    sess = get_session()

    try:
        del classifier # this is from global space - change this as you need
    except:
        pass

    print(gc.collect()) # if it does something you should see a number as output

    # use the same config as you used to create the session
    config = tf.compat.v1.ConfigProto()
    config.gpu_options.per_process_gpu_memory_fraction = 1
    config.gpu_options.visible_device_list = "0"
    set_session(tf.compat.v1.Session(config=config))
    gpu_options = tf.compat.v1.GPUOptions(allow_growth=True)
    sess = tf.compat.v1.Session(config=tf.compat.v1.ConfigProto(gpu_options=gpu_options))
    tf.compat.v1.keras.backend.set_session(sess)
    tf.compat.v1.enable_eager_execution()
#模型訓練用
def modeltrain(Xtrain,Ytrain,img_shape,savepath):
    batch_size=200
    no_epochs=500
    verbosity=2
    validation_split=0.3
    #classes指詞遷入向量大小
    classes=260
    momentum = 0.00
    actiname = 'tanh'
    # actiname = 'LeakyReLU'
#activation='linear'
#activation='tanh'
#activation='elu'

    #!!!!!
    input1 = Input(shape=(img_shape[1])) #260
    # input1 = Input(shape=(img_shape),batch_size=batch_size)
    # h = Dense(4096,activation=actiname)(input1)
    # h= BatchNormalization()(h)
    # # h = Dropout(0.10)(h)
    # h = Dense(2048,activation=actiname)(input1)
    # h= BatchNormalization()(h)
    # h = Dropout(0.30)(h)
    # h = Dense(1024,activation=actiname)(h)
    # h= BatchNormalization()(h)
    # h = Dropout(0.30)(h)

    #原本
    # h = Dense(4000,activation=actiname,kernel_regularizer=regularizers.l2(0.01))(input1)
    # h= BatchNormalization()(h)
    # h = Dropout(0.20)(h)
    # h = Dense(2500,activation=actiname,kernel_regularizer=regularizers.l2(0.01))(h)
    # h= BatchNormalization()(h)
    # h = Dropout(0.20)(h)
    h = Dense(4000,activation=actiname,kernel_regularizer=regularizers.l2(0.001))(input1)
    h= BatchNormalization()(h)
    h = Dropout(0.10)(h)
    h = Dense(2500,activation=actiname,kernel_regularizer=regularizers.l2(0.001))(h)
    h= BatchNormalization()(h)
    h = Dropout(0.10)(h)

    # h = Dense(128,activation=actiname)(h)
    # h= BatchNormalization()(h)
    # h = Dropout(0.30)(h)
    f1 = Dense(classes,activation='linear')(h)      
    model = Model(outputs=f1,inputs=[input1])
    model.summary()


    model.compile(loss='mean_squared_error',
              optimizer=keras.optimizers.Adagrad(learning_rate=lr,epsilon = epsilon),
              metrics=[['mean_squared_error']])
# optimizer=keras.optimizers.Adam(lr=lr,beta_1=beta_1,beta_2=beta_2,epsilon=epsilon,decay=decay)
# optimizer=keras.optimizers.SGD(lr=lr,decay=decay)
# optimizer=tf.keras.optimizers.Ftrl(lr=lr,decay=decay)
# optimizer=keras.optimizers.Adagrad(lr=lr)

    early_stop = EarlyStopping(monitor='mean_squared_error', patience=no_epochs, verbose=1,baseline=0.2, mode='max',restore_best_weights=False)
    savepoint_path = f"{savepath}MLP_{classes}_ver{ver}_{actiname}__flax{flax}_"
    savepoint_path = savepoint_path + "{epoch:02d}" + ".h5"
    checkpoint = ModelCheckpoint(savepoint_path,verbose=1,period=peroid,save_weights_only=False)
    #!!!!!
    model.fit(Xtrain,Ytrain
        ,batch_size=batch_size
        ,epochs=no_epochs,
        callbacks=[early_stop,checkpoint],
        verbose=verbosity,
        validation_split=validation_split)
    model.save(f'{savepath}MLP_{classes}_ver{ver}_complete_{actiname}_flax{flax}.h5')
#模型測試用(train_test_split測試) 沒有提供top-5準確率
def result(X,Y,savepath,model,testvecotr):
    label = []
    
    with open(testtxt, 'r', encoding='utf-8') as labels:
        for temp in labels:
            if "：" in temp:
                te = temp.split(".")
                te = te[1].split("：")
                label.append(te[0])
    total = len(Y)
    y_ture_L = []
    with alive_bar(total) as bar:
        for yt in Y:
            for k in range(1000):
                vtotal = 0
                # for v in range(classes):
                #     vc = yt[v] - testvecotr[k][v]
                #MSE公式
                vc = cul(testvecotr[k],yt)
                # vc =np.mean((yt - testvecotr[k])**2) 
                    # vtotal = vtotal + abs(vc)
                vtotal = vc
                if vtotal == 0:
                    y_ture_L.append(k)
                    break
            bar()
    model=tf.keras.models.load_model(f'{savepath}{model}')
    y_prob=model.predict(X) #預測機率
    y_prob_L = []
    
    total = len(y_prob)
    with alive_bar(total) as bar:
        for y in y_prob:
            min = 0
            min_lab = -1
            for k in range(1000):
                vtotal = 0
                # for v in range(classes):
                    # vc = y[v] - testvecotr[k][v] #距離
                vc = cul(testvecotr[k],yt)
                # vc = np.mean((testvecotr[k] - y)**2)  #MSE
                vtotal = vtotal + vc
                #改成MSE的計算公式
                #Top5正確率試試看(距離前5短)
                if min_lab == -1 :
                    min = vtotal
                    min_lab = k
                elif min > vtotal:
                    min = vtotal
                    min_lab = k
                bar.text("Sub_Progessing...{}%".format(k/10))
            y_prob_L.append(min_lab)
            bar()
    Accuracy=accuracy_score(y_ture_L,y_prob_L)    
    return Accuracy
    # print(Accuracy)
    # y_classes=np.argmax(y_prob,axis=1)
    # Y=np.argmax(Y,axis=1)
    # #print(y_prob)
    # #y_classes = [1 if v[0] >= 0.5 else 0 for v in y_prob]
    #   #  y_classes = np.array(y_prob)
    # #print(y_classes)
    # #print(Y)


    # matrix=confusion_matrix(Y,y_classes,labels=[0,1,2]) #混淆矩陣
    # print(matrix)
    # Report=classification_report(Y,y_classes,labels=[0,1,2],output_dict=True) #分類報告（含Precision,Recall）
    # Precision=Report['0']['precision']
    # Recall=Report['0']['recall']
    # Accuracy=accuracy_score(Y,y_classes)
    # print(f'Precision={Precision}')
    # print('Recall=',Recall)
    # print('Accuracy=',Accuracy)

    # wb=Workbook()

    # ws=wb.active


    # ws['F17']="Precision"
    # ws['F18']=round(Precision,4)
    # ws['G17']="Recall"
    # ws['G18']=round(Recall,4)
    # ws['H17']="Accuracy"
    # ws['H18']=round(Accuracy,4)
    # ws['I17']="F1_score"
    # ws['I18']=round(2*Recall*Precision/(Recall+Precision),4)

    # a ='evaluate'
    # col = ['A','B','C','D','E','F']
    # for i in range(len(matrix)):
    #     ws[f"A{i+2}"]=f"Predict{i}"
    #     ws[f"{col[i+1]}1"]=f"Real{i}"

    #     for y in range(len(matrix)):
    #         ws[f"{col[i+1]}{y+2}"]=matrix[i][y]

    # col = ['G','H','I']
    # ws['G1']= "訓練"
    # ws['G4']="測試"
    # for i in range(len(col)):
    #     ws[f"{col[i]}2"]=f"label={i}"
    #     ws[f"{col[i]}3"]= trdc[i]
    #     ws[f"{col[i]}5"]=f"label={i}"
    #     ws[f"{col[i]}6"]= tsdc[i]
    

    
    # cols=['A','B','C','F','G','H','I','J']
    # for col in cols:
    #     ws.column_dimensions[col].width=15
        
    # wb.save(f'{savepath}gait.xlsx')
    # print("result_done")


#train_test分割
#全部未標註圖片這台電腦的極限是 test_size = 0.4 0.39
x_tran,x_test,y_train,y_test = train_test_split(xvector, yvector, test_size=0.4)
temp = x_tran.shape
#!!!!!
# img_shape = [temp[1],temp[2]]
img_shape = [temp[0],temp[1]] #260
# reset_keras()
model_name = []
Acc_set = []
#模型訓練
modeltrain(x_tran,y_train,img_shape,savepath)

# testpath = savepath
#模型測試 會自動抓取testpath裡面所有的模型

def test_model():
    for model in os.listdir(testpath) :
        classes = int(model.split("_")[1])
        #因為圖片來源一樣所以classes不用指定
        xvector = np.load('D:/F111156114_Aliline/imgtoword/wordtovector/vector_MLP_renew/vx260_full_corp.npy',allow_pickle=True)
        #會自動抓模型採用的多大的詞遷入向量
        yvector = np.load('D:/F111156114_Aliline/imgtoword/wordtovector/vector_MLP_renew/v{}y_full_corp.npy'.format(classes))
        x_tran,x_test,y_train,y_test = train_test_split(xvector, yvector, test_size=0.04)
        testvecotr = np.load("D:/F111156114_Aliline/imgtoword/wordtovector/vectors_{}_ver2.npy".format(classes))
        acc = result(x_test,y_test,testpath,model,testvecotr)
        model_name.append(model)
        Acc_set.append(acc)
        print("model:{},Acc={}".format(model,acc))
    for k in range(len(model_name)):
        print("model:{},Acc={}".format(model_name[k],Acc_set[k]))

# test_model()
